"""
Issues #341 et #470 — Feedback Dashboard : filtres, ministère et export.

Teste les helpers purs (``src.ui.feedback_dashboard``) et, à la manière de
``test_rag_health_dashboard.py``, la requête/présentation de la page Streamlit
via son source (la page ne s'importe pas sans session Streamlit).
"""

from __future__ import annotations

from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.ui.feedback_dashboard import (
    BETA_END,
    BETA_START,
    MINISTRY_NOT_SET_LABEL,
    PERIOD_ALL,
    PERIOD_BETA,
    PERIOD_CUSTOM,
    PERIOD_LAST_MONTH,
    PERIOD_MODE_LABELS,
    PERIOD_MODE_OPTIONS,
    build_unified_feedback_export,
    current_paris_date,
    dataframe_to_xlsx_bytes,
    ministry_display_label,
    period_caption,
    previous_calendar_month,
    resolve_period,
    visible_available_groups,
)

REPO_ROOT = Path(__file__).resolve().parents[1]
DASHBOARD_PATH = REPO_ROOT / "apps" / "streamlit-ui" / "pages" / "03_Feedback_Dashboard.py"


def _dashboard_source() -> str:
    return DASHBOARD_PATH.read_text(encoding="utf-8")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Période — mode explicite, plage dynamique en « Tout », personnalisée figée
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


class TestResolvePeriod:
    def test_all_mode_has_no_frozen_bound(self):
        # « Tout » ne fige aucune date de fin : les feedbacks ajoutés après
        # l'ouverture de la vue restent visibles à chaque rechargement.
        assert resolve_period(PERIOD_ALL) is None

    def test_all_mode_ignores_stale_custom_range(self):
        stale = (date(2026, 1, 1), date(2026, 7, 1))
        assert resolve_period(PERIOD_ALL, stale) is None

    def test_beta_preset_is_fixed_window(self):
        assert resolve_period(PERIOD_BETA) == (BETA_START, BETA_END)
        assert resolve_period(PERIOD_BETA) == (date(2026, 1, 8), date(2026, 2, 6))

    def test_last_month_is_previous_calendar_month(self):
        assert resolve_period(PERIOD_LAST_MONTH, reference_date=date(2026, 8, 31)) == (date(2026, 7, 1), date(2026, 7, 31))

    def test_last_month_crosses_year_boundary(self):
        assert previous_calendar_month(date(2026, 1, 10)) == (date(2025, 12, 1), date(2025, 12, 31))

    def test_current_date_uses_paris_timezone_at_utc_month_boundary(self):
        utc_instant = datetime(2026, 8, 31, 22, 30, tzinfo=timezone.utc)
        assert current_paris_date(utc_instant) == date(2026, 9, 1)

    def test_custom_range_is_stable_across_reruns(self):
        chosen = (date(2026, 6, 1), date(2026, 6, 30))
        first = resolve_period(PERIOD_CUSTOM, chosen)
        second = resolve_period(PERIOD_CUSTOM, chosen)
        assert first == second == chosen

    def test_custom_incomplete_selection_keeps_last_applied_range(self):
        # st.date_input renvoie un tuple à 1 élément en cours de sélection :
        # la dernière plage complète reste appliquée, la vue ne s'élargit pas
        # silencieusement à « Tout ».
        last = (date(2026, 6, 1), date(2026, 6, 30))
        assert resolve_period(PERIOD_CUSTOM, (date(2026, 7, 1),), last) == last
        assert resolve_period(PERIOD_CUSTOM, None, last) == last

    def test_custom_complete_selection_overrides_last_applied(self):
        last = (date(2026, 6, 1), date(2026, 6, 30))
        chosen = (date(2026, 7, 1), date(2026, 7, 15))
        assert resolve_period(PERIOD_CUSTOM, chosen, last) == chosen

    def test_custom_incomplete_without_history_applies_no_bound(self):
        assert resolve_period(PERIOD_CUSTOM, (date(2026, 6, 1),)) is None
        assert resolve_period(PERIOD_CUSTOM, None) is None

    def test_every_mode_has_a_label(self):
        assert set(PERIOD_MODE_OPTIONS) == set(PERIOD_MODE_LABELS)


class TestPeriodCaption:
    def test_all_mode_shows_data_bounds(self):
        caption = period_caption(PERIOD_ALL, None, date(2026, 1, 8), date(2026, 7, 23))
        assert "Tout" in caption
        assert "08/01/2026" in caption
        assert "23/07/2026" in caption

    def test_all_mode_without_data(self):
        assert "Tout" in period_caption(PERIOD_ALL, None)

    def test_custom_mode_shows_applied_bounds(self):
        caption = period_caption(PERIOD_CUSTOM, (date(2026, 6, 1), date(2026, 6, 30)))
        assert "Personnalisée" in caption
        assert "01/06/2026" in caption
        assert "30/06/2026" in caption

    def test_beta_mode_labelled(self):
        caption = period_caption(PERIOD_BETA, (BETA_START, BETA_END))
        assert "Beta-test" in caption

    def test_last_month_mode_labelled(self):
        caption = period_caption(PERIOD_LAST_MONTH, (date(2026, 7, 1), date(2026, 7, 31)))
        assert "Mois dernier" in caption
        assert "01/07/2026" in caption
        assert "31/07/2026" in caption

    def test_custom_incomplete_is_not_announced_as_tout(self):
        caption = period_caption(PERIOD_CUSTOM, None, date(2026, 1, 8), date(2026, 7, 23))
        assert "Tout" not in caption
        assert "Personnalisée" in caption


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Ministère — libellé lisible, « Non renseigné » sans backfill heuristique
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


class TestMinistryDisplayLabel:
    def test_known_ids_resolve_to_catalog_labels(self):
        assert ministry_display_label("matte") == "MATTE"
        assert ministry_display_label("mso") == "MSO"
        assert ministry_display_label("mi") == "MI"
        assert ministry_display_label("masa") == "MASA"

    def test_historic_rows_without_value_show_non_renseigne(self):
        assert ministry_display_label(None) == MINISTRY_NOT_SET_LABEL
        assert ministry_display_label("") == MINISTRY_NOT_SET_LABEL
        assert ministry_display_label(float("nan")) == MINISTRY_NOT_SET_LABEL

    def test_unknown_id_passes_through_without_guessing(self):
        assert ministry_display_label("eval_all_ministries") == "eval_all_ministries"


class TestVisibleAvailableGroups:
    def test_hidden_groups_are_excluded_from_dashboard_filter(self):
        configured = [
            {"slug": "visible-group", "visible": True},
            {"slug": "hidden-group", "visible": False},
        ]

        assert visible_available_groups(["hidden-group", "historical-group", "visible-group"], configured) == [
            "historical-group",
            "visible-group",
        ]

    def test_missing_visible_flag_defaults_to_visible(self):
        assert visible_available_groups(["legacy-group"], [{"slug": "legacy-group"}]) == ["legacy-group"]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Issue #470 — export Excel unifié, jointure gauche et dédoublonnage
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


def _questions_for_export() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "ts": "2026-07-17T10:00:00Z",
                "turn_id": "with-feedback",
                "session_id": "session-1",
                "user_group": "betatest-jan26",
                "selected_ministry": "matte",
                "question": "Question évaluée",
                "answer": "Réponse évaluée",
                "theme": "conges",
                "rag_version": "v3",
                "chunk_selection_mode": "section",
                "dist_after_rerank": "{}",
                "total_time_ms": 1200,
            },
            {
                "ts": "2026-07-17T11:00:00Z",
                "turn_id": "without-feedback",
                "session_id": "session-2",
                "user_group": "betatest-jan26",
                "selected_ministry": None,
                "question": "Question sans évaluation",
                "answer": "Réponse sans évaluation",
                "theme": "remuneration",
                "rag_version": "v3",
                "chunk_selection_mode": "section",
                "dist_after_rerank": "{}",
                "total_time_ms": 900,
            },
        ]
    )


def _feedback_for_export() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "id": 42,
                "ts": "2026-07-17T10:05:00Z",
                "turn_id": "with-feedback",
                "stars": 4,
                "helpful": True,
                "reasons_positive": "Sources claires",
                "reasons_negative": "",
                "comment": "Très utile",
                "theme": "conges",
                "beta_scope": "Oui",
                "error_category": "",
                "ai_reason": "",
            }
        ]
    )


class TestUnifiedFeedbackExport:
    def test_keeps_every_question_and_enriches_only_matching_turn(self):
        export = build_unified_feedback_export(_questions_for_export(), _feedback_for_export())

        assert export["turn_id"].tolist() == ["with-feedback", "without-feedback"]
        evaluated = export.loc[export["turn_id"] == "with-feedback"].iloc[0]
        assert evaluated["Note (1-5)"] == 5
        assert bool(evaluated["Utile"]) is True
        assert evaluated["Commentaire"] == "Très utile"
        assert evaluated["Date"] == "17/07/2026 12:00"
        assert evaluated["Date du feedback"] == "17/07/2026 12:05"
        assert evaluated["Groupe utilisateur"] == "betatest-jan26"
        assert evaluated["Ministère"] == "MATTE"

        unanswered = export.loc[export["turn_id"] == "without-feedback"].iloc[0]
        assert unanswered["Question"] == "Question sans évaluation"
        assert pd.isna(unanswered["feedback_id"])
        assert pd.isna(unanswered["Note (1-5)"])
        assert pd.isna(unanswered["Commentaire"])
        assert unanswered["Ministère"] == MINISTRY_NOT_SET_LABEL

    def test_keeps_all_questions_when_no_feedback_is_eligible(self):
        export = build_unified_feedback_export(_questions_for_export(), pd.DataFrame())

        assert export["turn_id"].tolist() == ["with-feedback", "without-feedback"]
        assert export["feedback_id"].isna().all()

    def test_empty_question_set_still_has_the_stable_export_schema(self):
        export = build_unified_feedback_export(pd.DataFrame(), _feedback_for_export())

        assert export.empty
        assert ["Date", "Groupe utilisateur", "Ministère", "Question", "Réponse"] == export.columns[:5].tolist()

    def test_feedback_theme_fills_missing_historic_question_theme(self):
        questions = _questions_for_export().iloc[[0]].copy()
        questions["theme"] = None

        export = build_unified_feedback_export(questions, _feedback_for_export())

        assert export.loc[0, "Thème"] == "conges"

    def test_preserves_existing_question_technical_metadata(self):
        export = build_unified_feedback_export(_questions_for_export(), _feedback_for_export())

        assert export.loc[0, "session_id"] == "session-1"
        assert export.loc[0, "rag_version"] == "v3"
        assert export.loc[0, "chunk_selection_mode"] == "section"
        assert export.loc[0, "dist_after_rerank"] == "{}"
        assert export.loc[0, "total_time_ms"] == 1200

    def test_latest_feedback_wins_then_highest_id_breaks_timestamp_tie(self):
        questions = _questions_for_export().iloc[[0]]
        feedbacks = pd.DataFrame(
            [
                {"id": 99, "turn_id": "with-feedback", "ts": "2026-07-17T09:00:00Z", "comment": "ancien"},
                {"id": 7, "turn_id": "with-feedback", "ts": "2026-07-17T10:00:00Z", "comment": "même date, petit id"},
                {"id": 8, "turn_id": "with-feedback", "ts": "2026-07-17T10:00:00Z", "comment": "retenu"},
            ]
        )

        export = build_unified_feedback_export(questions, feedbacks)

        assert len(export) == 1
        assert export.loc[0, "feedback_id"] == 8
        assert export.loc[0, "Commentaire"] == "retenu"

    def test_serialized_workbook_is_readable(self):
        export = build_unified_feedback_export(_questions_for_export(), _feedback_for_export())

        workbook = load_workbook(BytesIO(dataframe_to_xlsx_bytes(export, "Questions et feedbacks")), read_only=True)
        worksheet = workbook["Questions et feedbacks"]

        assert worksheet.max_row == 3
        assert [cell.value for cell in next(worksheet.iter_rows(max_row=1))] == export.columns.tolist()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Page Streamlit — requête et présentation
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


class TestDashboardSource:
    def test_feedback_query_selects_ministry_via_run_join(self):
        assert "r.selected_ministry" in _dashboard_source()

    def test_questions_query_selects_ministry(self):
        source = _dashboard_source()
        assert source.count("selected_ministry") >= 2

    def test_questions_query_selects_theme_for_unanswered_questions(self):
        assert "v3_detected_theme AS theme" in _dashboard_source()

    def test_ministry_column_displayed_and_exported(self):
        source = _dashboard_source()
        assert "ministere_display" in source
        assert "🏛️ Ministère" in source

    def test_no_frozen_date_range_left_in_session_state(self):
        # L'ancien mécanisme figeait la fin de période dans fb_date_range.
        assert "fb_date_range" not in _dashboard_source()

    def test_period_mode_wired_to_helpers(self):
        source = _dashboard_source()
        assert "resolve_period(" in source
        assert "period_caption(" in source
        assert "fb_period_mode" in source

    def test_group_filter_uses_admin_visibility(self):
        source = _dashboard_source()
        assert "visible_available_groups(" in source
        assert "list_groups()" in source

    def test_grid_receives_datetimes_not_preformatted_strings(self):
        # Une chaîne "17/07/2026 …" passée à une DatetimeColumn est re-parsée
        # mois/jour par la grille (10/07 affiché 7 octobre) : la grille doit
        # recevoir des datetimes ; le strftime est réservé à l'export Excel.
        source = _dashboard_source()
        assert 'feedback_display_df["ts"].dt.tz_localize(None)' in source
        assert 'feedback_display_df["ts"].dt.strftime' not in source

    def test_single_unified_excel_export_replaces_both_buttons(self):
        source = _dashboard_source()
        assert source.count("st.download_button(") == 1
        assert "Export questions, réponses et feedbacks (Excel)" in source
        assert "Export Feedbacks (Excel)" not in source
        assert "Export Questions (Excel)" not in source

    def test_export_is_built_after_theme_and_quality_filters(self):
        source = _dashboard_source()
        export_position = source.index("build_unified_feedback_export(df_q, df_f)")
        assert source.index("df_q = df_q[question_theme_mask]") < export_position
        assert source.index("df_f = df_f[~exclude_mask]") < export_position

    def test_duplicate_feedback_rule_is_documented_in_dashboard(self):
        source = _dashboard_source()
        assert "feedbacks multiples" in source
        assert "l’identifiant le plus élevé" in source

    def test_candidate_cut_has_its_own_aggregate_bucket(self):
        source = _dashboard_source()
        assert 'elif cat == "candidate_cut":' in source
        assert '"📉 Agrégation (coupe candidats)": "#19D3F3"' in source
